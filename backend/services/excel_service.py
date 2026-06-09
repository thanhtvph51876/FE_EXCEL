from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Dict, Iterable, List

import openpyxl
import pandas as pd


FORMULA_INJECTION_PREFIXES = ("=", "+", "-", "@")


@dataclass
class ParsedWorkbook:
    headers: List[str]
    rows: List[List[str]]

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def col_count(self) -> int:
        return len(self.headers)

    @property
    def preview_rows(self) -> List[List[str]]:
        return self.rows[:100]


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def escape_formula_injection(value: Any) -> str:
    text = stringify(value)
    if text.startswith(FORMULA_INJECTION_PREFIXES):
        return "'" + text
    return text


def escape_formula_row(row: Iterable[Any]) -> List[str]:
    return [escape_formula_injection(value) for value in row]


def format_file_size(size_bytes: int) -> str:
    if size_bytes < 1024:
        return f"{size_bytes} B"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes / 1024 / 1024:.2f} MB"


def normalize_headers(raw_headers: Iterable[Any]) -> List[str]:
    headers = []
    seen: Dict[str, int] = {}
    for index, value in enumerate(raw_headers):
        name = stringify(value) or f"Cột {index + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        headers.append(name)
    return headers


def _rows_from_dataframe(df: pd.DataFrame) -> ParsedWorkbook:
    df = df.fillna("")
    headers = normalize_headers(df.columns.tolist())
    rows = [[stringify(value) for value in row] for row in df.to_numpy().tolist()]
    return ParsedWorkbook(headers=headers, rows=rows)


def _parse_csv(content: bytes) -> ParsedWorkbook:
    for encoding in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = content.decode("utf-8", errors="replace")

    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    df = pd.read_csv(io.StringIO(text), dtype=str, keep_default_na=False, sep=delimiter)
    return _rows_from_dataframe(df)


def _parse_xlsx(content: bytes) -> ParsedWorkbook:
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    row_iter = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(row_iter)
    except StopIteration:
        return ParsedWorkbook(headers=[], rows=[])
    headers = normalize_headers(raw_headers)
    rows = []
    for row in row_iter:
        normalized = [stringify(value) for value in row[: len(headers)]]
        while len(normalized) < len(headers):
            normalized.append("")
        if any(cell != "" for cell in normalized):
            rows.append(normalized)
    workbook.close()
    return ParsedWorkbook(headers=headers, rows=rows)


def _parse_xls(content: bytes) -> ParsedWorkbook:
    df = pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False, engine="xlrd")
    return _rows_from_dataframe(df)


def parse_workbook(filename: str, content: bytes) -> ParsedWorkbook:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _parse_csv(content)
    if lower.endswith(".xlsx"):
        return _parse_xlsx(content)
    if lower.endswith(".xls"):
        return _parse_xls(content)
    raise ValueError("Định dạng file không hỗ trợ.")


def _looks_like_number(value: str) -> bool:
    cleaned = re.sub(r"[,\s₫đ]", "", value)
    return cleaned != "" and re.fullmatch(r"-?\d+(\.\d+)?", cleaned) is not None


def _looks_like_date(value: str) -> bool:
    if not value:
        return False
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            datetime.strptime(value, fmt)
            return True
        except ValueError:
            continue
    return False


def build_statistics(headers: List[str], rows: List[List[str]], total_rows: int) -> Dict[str, Any]:
    stats = {"totalRows": total_rows, "totalCols": len(headers), "missingValues": 0, "duplicateRows": 0, "columns": []}
    seen = set()
    for row in rows:
        row_key = "|".join(row)
        if row_key in seen:
            stats["duplicateRows"] += 1
        else:
            seen.add(row_key)

    for col_index, header in enumerate(headers):
        empty = 0
        number_count = 0
        date_count = 0
        value_counts: Dict[str, int] = {}
        for row in rows:
            value = row[col_index] if col_index < len(row) else ""
            if not value:
                empty += 1
                continue
            if _looks_like_number(value):
                number_count += 1
            if _looks_like_date(value):
                date_count += 1
            value_counts[value] = value_counts.get(value, 0) + 1

        stats["missingValues"] += empty
        non_empty = max(1, len(rows) - empty)
        if date_count / non_empty >= 0.6:
            col_type = "Ngày"
        elif number_count / non_empty >= 0.6:
            col_type = "Số"
        else:
            col_type = "Văn bản"
        top_values = sorted(value_counts.items(), key=lambda item: item[1], reverse=True)[:3]
        stats["columns"].append(
            {
                "name": header,
                "type": col_type,
                "missingCount": empty,
                "topValues": ", ".join(f"{value} ({count} lần)" for value, count in top_values) or "Không có",
            }
        )
    return stats


def to_number(value: Any) -> float:
    cleaned = re.sub(r"[^\d,\.\-]", "", stringify(value))
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", "")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def find_quality_errors(headers: List[str], rows: List[List[str]], max_errors: int = 200) -> List[Dict[str, Any]]:
    errors: List[Dict[str, Any]] = []
    lowercase_headers = [header.lower() for header in headers]

    def add(row_index: int, col_index: int, value: str, issue: str) -> None:
        if len(errors) >= max_errors:
            return
        errors.append({"row": row_index + 2, "column": headers[col_index], "value": value, "issue": issue})

    email_indexes = [idx for idx, header in enumerate(lowercase_headers) if "email" in header or "thư" in header]
    date_indexes = [idx for idx, header in enumerate(lowercase_headers) if "ngày" in header or "date" in header]
    amount_indexes = [
        idx
        for idx, header in enumerate(lowercase_headers)
        if any(key in header for key in ("tiền", "amount", "price", "doanh thu", "giá", "số lượng", "quantity", "qty"))
    ]
    email_regex = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")

    for row_index, row in enumerate(rows[:100]):
        for col_index, _ in enumerate(headers):
            value = row[col_index] if col_index < len(row) else ""
            if value == "":
                add(row_index, col_index, value, "Ô trống hoặc thiếu dữ liệu")
                continue
            if col_index in email_indexes and not email_regex.match(value.strip()):
                add(row_index, col_index, value, "Lỗi định dạng Email thiếu hoặc sai ký tự '@'")
            if col_index in date_indexes and not _looks_like_date(value.strip()):
                add(row_index, col_index, value, "Ngày không hợp lệ hoặc sai định dạng")
            if col_index in amount_indexes and to_number(value) < 0:
                add(row_index, col_index, value, "Giá trị âm bất thường")
    return errors


def clean_value(value: str, rule: str) -> str:
    if rule == "trim":
        return re.sub(r"\s+", " ", value).strip()
    if rule == "upper":
        return value.upper()
    if rule == "lower":
        return value.lower()
    if rule == "phone":
        cleaned = re.sub(r"[\s\-\(\)]", "", value)
        cleaned = re.sub(r"^\+84", "0", cleaned)
        return cleaned if cleaned.startswith("0") else f"0{cleaned}"
    if rule == "email":
        return value.strip().lower()
    if rule == "name":
        return " ".join(part.capitalize() for part in value.strip().split())
    return value
