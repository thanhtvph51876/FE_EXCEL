from __future__ import annotations

import json
import sys
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path

BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from config import settings
from dependencies import get_db
from services.storage_service import StorageService


CATEGORIES = [
    ("finance", "Tai chinh", "💰", "green", ["Ngan sach", "Dong tien", "Cong no", "Loi nhuan", "Chi phi"]),
    ("sales", "Ban hang", "📈", "blue", ["Pipeline", "Doanh so", "Khach hang", "Bao gia", "Hoa don"]),
    ("hr", "Nhan su", "👥", "purple", ["Cham cong", "Bang luong", "Tuyen dung", "Dao tao", "Danh gia"]),
    ("operations", "Van hanh", "⚙️", "orange", ["Ton kho", "Mua hang", "Ke hoach", "Chat luong", "Bao tri"]),
    ("project", "Du an", "📋", "teal", ["Gantt", "Task", "Rui ro", "Nguon luc", "Nghiem thu"]),
    ("marketing", "Marketing", "🎯", "pink", ["Campaign", "Content", "KPI", "Leads", "Social"]),
    ("admin", "Hanh chinh", "🏢", "slate", ["Tai san", "Hop dong", "Cong van", "Lich hop", "Van phong"]),
    ("analytics", "Bao cao", "📊", "indigo", ["Dashboard", "KPI", "Phan tich", "Doi soat", "Tong hop"]),
    ("education", "Giao duc", "🎓", "cyan", ["Lop hoc", "Diem so", "Hoc phi", "Lich hoc", "Giao vien"]),
    ("personal", "Ca nhan", "✅", "amber", ["Ke hoach", "Thu chi", "Muc tieu", "Thoi quen", "Tai chinh"]),
]

SHEET_COLUMNS = {
    "finance": ["Ngay", "Bo phan", "Hang muc", "Ngan sach", "Thuc te", "Chenh lech", "Ty le", "Trang thai", "Nguoi phu trach", "Ghi chu"],
    "sales": ["Ngay", "Khach hang", "San pham", "Kenh", "So luong", "Don gia", "Doanh thu", "Ty le chot", "Trang thai", "Phu trach"],
    "hr": ["Ma NV", "Ho ten", "Bo phan", "Chuc danh", "Ngay cong", "Luong co ban", "Phu cap", "Khau tru", "Tong cong", "Trang thai"],
    "operations": ["Ma hang", "Ten hang", "Nhom", "Ton dau", "Nhap", "Xuat", "Ton cuoi", "Dinh muc", "Canh bao", "Phu trach"],
    "project": ["Cong viec", "Hang muc", "Phu trach", "Bat dau", "Ket thuc", "Tien do", "Ngan sach", "Thuc te", "Rui ro", "Trang thai"],
    "marketing": ["Ngay", "Kenh", "Chien dich", "Chi phi", "Hien thi", "Clicks", "Leads", "Chuyen doi", "CPA", "ROI"],
    "admin": ["Ma muc", "Noi dung", "Bo phan", "Ngay", "Han xu ly", "Nguoi phu trach", "Muc do", "Chi phi", "Trang thai", "Ghi chu"],
    "analytics": ["Ky bao cao", "Chi so", "Nhom", "Muc tieu", "Thuc hien", "Chenh lech", "Ty le hoan thanh", "Xu huong", "Nguon du lieu", "Ghi chu"],
    "education": ["Ma lop", "Ten lop", "Giao vien", "Hoc vien", "Diem TB", "Ty le tham gia", "Hoc phi", "Da thu", "Con lai", "Trang thai"],
    "personal": ["Ngay", "Nhom", "Noi dung", "Muc tieu", "Thuc te", "Chenh lech", "Uu tien", "Trang thai", "Han hoan thanh", "Ghi chu"],
}

STATUS_VALUES = ["Moi", "Dang xu ly", "Dung tien do", "Can chu y", "Hoan tat"]
OWNERS = ["Nguyen An", "Tran Binh", "Le Chi", "Pham Dung", "Hoang Linh", "Do Mai"]
DEPARTMENTS = ["Kinh doanh", "Tai chinh", "Nhan su", "Van hanh", "Marketing", "Ban giam doc"]
DATA_START_ROW = 5
DATA_END_ROW = 64
HEADER_ROW = 4


def slugify(value: str) -> str:
    normalized = (
        value.lower()
        .replace(" ", "-")
        .replace("/", "-")
        .replace("&", "and")
    )
    return "".join(ch for ch in normalized if ch.isalnum() or ch == "-").strip("-")


def build_catalog() -> list[dict]:
    templates: list[dict] = []
    counter = 1
    for category_id, category_name, icon, color, topics in CATEGORIES:
        for topic_index in range(10):
            topic = topics[topic_index % len(topics)]
            variant = topic_index // len(topics) + 1
            name = f"{topic} {category_name} chuyen nghiep {variant}"
            template_id = f"tpl-{category_id}-{topic_index + 1:02d}"
            templates.append(
                {
                    "id": template_id,
                    "name": name,
                    "category": category_name,
                    "description": f"Mau Excel that cho {topic.lower()} trong nhom {category_name.lower()}, co sheet nhap lieu, tong hop KPI va huong dan su dung.",
                    "file": f"{slugify(name)}.xlsx",
                    "icon": icon,
                    "color": color,
                    "topic": topic,
                    "category_id": category_id,
                    "created_at": datetime.now(timezone.utc) - timedelta(minutes=counter),
                }
            )
            counter += 1
    return templates


def style_title(ws, title: str, subtitle: str, width: int = 10) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)


def style_data_sheet(ws, title: str, category_id: str) -> None:
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    columns = SHEET_COLUMNS[category_id]
    style_title(ws, title, "Nhap du lieu tai bang ben duoi. Cac cot tinh toan da co cong thuc mau.", len(columns))
    for col_index, column_name in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_index, value=column_name)
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = section_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        for col in range(1, len(columns) + 1):
            column = columns[col - 1]
            cell = ws.cell(row=row, column=col, value=sample_value(column, row, category_id))
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if any(key in column for key in ["Ngay", "Bat dau", "Ket thuc", "Han"]):
                cell.number_format = "dd/mm/yyyy"
            elif any(key in column for key in ["Ngan sach", "Thuc te", "Chi phi", "Doanh thu", "Luong", "Phu cap", "Khau tru", "Tong cong", "Hoc phi", "Da thu", "Con lai", "Don gia"]):
                cell.number_format = '#,##0 "VND"'
            elif any(key in column for key in ["Ty le", "Tien do", "ROI", "CPA"]):
                cell.number_format = "0.0%"

    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16 if col <= 3 else 14
    ws.freeze_panes = "A5"
    table_ref = f"A{HEADER_ROW}:{get_column_letter(len(columns))}{DATA_END_ROW}"
    table = Table(displayName=f"tbl_{category_id}_{ws.title.replace(' ', '_')[:12]}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    ws.add_table(table)
    ws.auto_filter.ref = table_ref

    status_col = next((i + 1 for i, name in enumerate(columns) if name == "Trang thai"), None)
    warning_col = next((i + 1 for i, name in enumerate(columns) if name == "Canh bao"), None)
    owner_col = next((i + 1 for i, name in enumerate(columns) if "Phu trach" in name or "Nguoi phu trach" in name), None)
    department_col = next((i + 1 for i, name in enumerate(columns) if name in {"Bo phan", "Nhom"}), None)
    level_col = next((i + 1 for i, name in enumerate(columns) if name in {"Uu tien", "Rui ro", "Muc do"}), None)
    channel_col = next((i + 1 for i, name in enumerate(columns) if name == "Kenh"), None)
    product_col = next((i + 1 for i, name in enumerate(columns) if name == "San pham"), None)
    if status_col:
        dv = DataValidation(type="list", formula1='"' + ",".join(STATUS_VALUES) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(status_col)}{DATA_START_ROW}:{get_column_letter(status_col)}200")
    if owner_col:
        dv = DataValidation(type="list", formula1='"' + ",".join(OWNERS) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(owner_col)}{DATA_START_ROW}:{get_column_letter(owner_col)}200")
    if department_col:
        dv = DataValidation(type="list", formula1='"' + ",".join(DEPARTMENTS) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(department_col)}{DATA_START_ROW}:{get_column_letter(department_col)}200")
    if level_col:
        dv = DataValidation(type="list", formula1='"Thap,Trung binh,Cao"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(level_col)}{DATA_START_ROW}:{get_column_letter(level_col)}200")
    if channel_col:
        dv = DataValidation(type="list", formula1='"Website,Facebook,Google,Email,Doi tac"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(channel_col)}{DATA_START_ROW}:{get_column_letter(channel_col)}200")
    if product_col:
        dv = DataValidation(type="list", formula1='"Goi Pro,Goi Business,Tu van,Bao tri"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(product_col)}{DATA_START_ROW}:{get_column_letter(product_col)}200")

    percent_col = next((i + 1 for i, name in enumerate(columns) if name in {"Ty le", "Ty le chot", "Tien do", "Ty le tham gia", "Ty le hoan thanh", "ROI"}), None)
    if percent_col:
        col_letter = get_column_letter(percent_col)
        ws.conditional_formatting.add(
            f"{col_letter}{DATA_START_ROW}:{col_letter}{DATA_END_ROW}",
            CellIsRule(operator="lessThan", formula=["0.7"], fill=PatternFill("solid", fgColor="FEE2E2")),
        )
        ws.conditional_formatting.add(
            f"{col_letter}{DATA_START_ROW}:{col_letter}{DATA_END_ROW}",
            CellIsRule(operator="greaterThanOrEqual", formula=["0.9"], fill=PatternFill("solid", fgColor="DCFCE7")),
        )
    if status_col:
        row_range = f"A{DATA_START_ROW}:{get_column_letter(len(columns))}{DATA_END_ROW}"
        status_letter = get_column_letter(status_col)
        ws.conditional_formatting.add(
            row_range,
            FormulaRule(formula=[f'${status_letter}{DATA_START_ROW}="Can chu y"'], fill=PatternFill("solid", fgColor="FEF3C7")),
        )
    if warning_col:
        row_range = f"A{DATA_START_ROW}:{get_column_letter(len(columns))}{DATA_END_ROW}"
        warning_letter = get_column_letter(warning_col)
        ws.conditional_formatting.add(
            row_range,
            FormulaRule(formula=[f'${warning_letter}{DATA_START_ROW}="Can nhap them"'], fill=PatternFill("solid", fgColor="FEF3C7")),
        )


def sample_value(column: str, row: int, category_id: str):
    n = row - 4
    if "Ngay" in column or "Bat dau" in column or "Han" in column:
        return datetime(2026, min(12, (n % 12) + 1), min(28, n)).date()
    if "Ket thuc" in column:
        return datetime(2026, min(12, (n % 12) + 1), min(28, n + 5)).date()
    if column in {"Bo phan", "Nhom"}:
        return DEPARTMENTS[n % len(DEPARTMENTS)]
    if any(key in column for key in ["Ngan sach", "Thuc te", "Chi phi", "Doanh thu", "Luong", "Phu cap", "Khau tru", "Tong cong", "Hoc phi", "Da thu", "Con lai", "Don gia"]):
        return 750_000 * (n + 2)
    if any(key in column for key in ["So luong", "Ton", "Nhap", "Xuat", "Leads", "Hoc vien", "Ngay cong", "Clicks", "Hien thi"]):
        return n * 7 + 12
    if any(key in column for key in ["Ty le", "Tien do", "ROI", "Chuyen doi", "Diem TB", "Ty le tham gia", "Ty le hoan thanh"]):
        return min(0.98, round(0.45 + (n % 10) / 20, 2))
    if "Trang thai" in column:
        return STATUS_VALUES[n % len(STATUS_VALUES)]
    if "Phu trach" in column or "Nguoi phu trach" in column or "Giao vien" in column:
        return OWNERS[n % len(OWNERS)]
    if "Khach hang" in column:
        return f"Cong ty {chr(64 + (n % 20) + 1)}"
    if "San pham" in column:
        return ["Goi Pro", "Goi Business", "Tu van", "Bao tri"][n % 4]
    if "Kenh" in column:
        return ["Website", "Facebook", "Google", "Email", "Doi tac"][n % 5]
    if "Muc do" in column or "Rui ro" in column or "Uu tien" in column:
        return ["Thap", "Trung binh", "Cao"][n % 3]
    if "Ky bao cao" in column:
        return f"2026-{(n % 12) + 1:02d}"
    return f"{column} {n:02d}"


def add_formula_columns(ws, category_id: str) -> None:
    columns = SHEET_COLUMNS[category_id]
    col_map = {name: i + 1 for i, name in enumerate(columns)}
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        if "Chenh lech" in col_map and "Ngan sach" in col_map and "Thuc te" in col_map:
            ws.cell(row=row, column=col_map["Chenh lech"], value=f"=E{row}-D{row}")
        if "Ty le" in col_map and "Thuc te" in col_map and "Ngan sach" in col_map:
            ws.cell(row=row, column=col_map["Ty le"], value=f"=IFERROR(E{row}/D{row},0)")
        if "Doanh thu" in col_map and "So luong" in col_map and "Don gia" in col_map:
            ws.cell(row=row, column=col_map["Doanh thu"], value=f"=E{row}*F{row}")
        if "Tong cong" in col_map and "Luong co ban" in col_map:
            ws.cell(row=row, column=col_map["Tong cong"], value=f"=F{row}+G{row}-H{row}")
        if "Ton cuoi" in col_map and "Ton dau" in col_map:
            ws.cell(row=row, column=col_map["Ton cuoi"], value=f"=D{row}+E{row}-F{row}")
        if "Canh bao" in col_map and "Ton cuoi" in col_map and "Dinh muc" in col_map:
            ws.cell(row=row, column=col_map["Canh bao"], value=f'=IF(G{row}<H{row},"Can nhap them","OK")')
        if "CPA" in col_map and "Chi phi" in col_map and "Leads" in col_map:
            ws.cell(row=row, column=col_map["CPA"], value=f"=IFERROR(D{row}/G{row},0)")
        if "ROI" in col_map and "Chi phi" in col_map and "Chuyen doi" in col_map:
            ws.cell(row=row, column=col_map["ROI"], value=f"=IFERROR((H{row}*500000-D{row})/D{row},0)")
        if "Con lai" in col_map and "Hoc phi" in col_map and "Da thu" in col_map:
            ws.cell(row=row, column=col_map["Con lai"], value=f"=G{row}-H{row}")


def dashboard_sheet(wb: Workbook, template: dict, category_id: str) -> None:
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    style_title(ws, f"Dashboard - {template['name']}", "Tong quan KPI tu sheet Nhap lieu", 10)
    kpis = [
        ("Tong dong", f"=COUNTA('Nhap lieu'!A{DATA_START_ROW}:A{DATA_END_ROW})"),
        ("Tong ke hoach", f"=SUM('Nhap lieu'!D{DATA_START_ROW}:D{DATA_END_ROW})"),
        ("Tong thuc hien", f"=SUM('Nhap lieu'!E{DATA_START_ROW}:E{DATA_END_ROW})"),
        ("Ty le TB", f"=AVERAGEIF('Nhap lieu'!G{DATA_START_ROW}:G{DATA_END_ROW},\">0\",'Nhap lieu'!G{DATA_START_ROW}:G{DATA_END_ROW})"),
    ]
    for idx, (label, formula) in enumerate(kpis, start=1):
        col = 1 + (idx - 1) * 2
        ws.cell(row=4, column=col, value=label)
        ws.cell(row=5, column=col, value=formula)
        ws.cell(row=4, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=4, column=col).fill = PatternFill("solid", fgColor="107C41")
        ws.cell(row=5, column=col).font = Font(size=14, bold=True)
        ws.cell(row=5, column=col).number_format = '#,##0'
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
    ws["A8"] = "Bang tom tat uu tien"
    ws["A8"].font = Font(bold=True, size=12)
    for col, value in enumerate(["Hang muc", "Ke hoach", "Thuc te", "Trang thai"], start=1):
        ws.cell(row=10, column=col, value=value)
    for row in range(11, 17):
        src = row - 6
        ws.cell(row=row, column=1, value=f"='Nhap lieu'!C{src}")
        ws.cell(row=row, column=2, value=f"='Nhap lieu'!D{src}")
        ws.cell(row=row, column=3, value=f"='Nhap lieu'!E{src}")
        ws.cell(row=row, column=4, value=f"='Nhap lieu'!H{src}")
    for cell in ws[10][0:4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 16
    chart = BarChart()
    chart.title = "Ke hoach vs Thuc te"
    chart.y_axis.title = "Gia tri"
    chart.x_axis.title = "Hang muc"
    data = Reference(ws, min_col=2, max_col=3, min_row=10, max_row=16)
    cats = Reference(ws, min_col=1, min_row=11, max_row=16)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, "F9")

    ws["A20"] = "Xu huong 12 muc dau"
    ws["A20"].font = Font(bold=True, size=12)
    for col, value in enumerate(["Muc", "Ke hoach", "Thuc hien"], start=1):
        ws.cell(row=22, column=col, value=value)
    for row in range(23, 35):
        src = row - 18
        ws.cell(row=row, column=1, value=f"='Nhap lieu'!C{src}")
        ws.cell(row=row, column=2, value=f"='Nhap lieu'!D{src}")
        ws.cell(row=row, column=3, value=f"='Nhap lieu'!E{src}")
    for cell in ws[22][0:3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    line = LineChart()
    line.title = "Xu huong ke hoach/thuc hien"
    line.y_axis.title = "Gia tri"
    line.x_axis.title = "Muc"
    line.add_data(Reference(ws, min_col=2, max_col=3, min_row=22, max_row=34), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=1, min_row=23, max_row=34))
    line.height = 8
    line.width = 16
    ws.add_chart(line, "F24")


def control_panel_sheet(wb: Workbook, template: dict) -> None:
    ws = wb.create_sheet("Control Panel")
    ws.sheet_view.showGridLines = False
    style_title(ws, "Control Panel", "Thiet lap tham so, muc tieu va thong tin quan tri template.", 6)
    rows = [
        ("Ten template", template["name"], "Thong tin nhan dien"),
        ("Nhom nghiep vu", template["category"], "Dung de loc va phan quyen"),
        ("Nam ke hoach", 2026, "Co the doi theo nam bao cao"),
        ("Don vi tien te", "VND", "Dung trong dinh dang tien"),
        ("Nguong canh bao", 0.7, "Ty le thap hon nguong se can chu y"),
        ("Muc tieu premium", 0.9, "Ty le dat ky vong"),
        ("Nguoi so huu", OWNERS[0], "Nguoi chiu trach nhiem cap nhat"),
        ("Phien ban", "2.0-premium", "Quan ly version noi bo"),
    ]
    ws.append([])
    ws.append(["Tham so", "Gia tri", "Ghi chu"])
    for row in rows:
        ws.append(list(row))
    for cell in ws[4][0:3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws["B8"].number_format = "0%"
    ws["B9"].number_format = "0%"
    for col, width in {"A": 24, "B": 24, "C": 58}.items():
        ws.column_dimensions[col].width = width


def data_dictionary_sheet(wb: Workbook, category_id: str) -> None:
    ws = wb.create_sheet("Tu dien du lieu")
    style_title(ws, "Tu dien du lieu", "Mo ta cot, kieu du lieu, muc bat buoc va cong thuc lien quan.", 6)
    headers = ["Cot", "Kieu du lieu", "Bat buoc", "Nguon/Validation", "Cong thuc/Ghi chu", "Dung cho AI"]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=value)
    for idx, column in enumerate(SHEET_COLUMNS[category_id], start=5):
        data_type = "Tien te" if any(key in column for key in ["Ngan sach", "Thuc te", "Chi phi", "Doanh thu", "Luong", "Hoc phi"]) else "Ngay" if any(key in column for key in ["Ngay", "Bat dau", "Ket thuc", "Han"]) else "Phan tram" if any(key in column for key in ["Ty le", "Tien do", "ROI"]) else "Van ban"
        validation = "Dropdown" if column in {"Trang thai", "Nguoi phu trach", "Phu trach", "Bo phan", "Nhom"} else "Nhap tay"
        note = "Cot tinh tu dong" if column in {"Chenh lech", "Ty le", "Doanh thu", "Tong cong", "Ton cuoi", "Canh bao", "CPA", "ROI", "Con lai"} else "Nguoi dung cap nhat"
        ws.append([column, data_type, "Co", validation, note, "Co"])
    for cell in ws[4][0:6]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    widths = [20, 16, 12, 20, 34, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    table_ref = f"A4:F{4 + len(SHEET_COLUMNS[category_id])}"
    table = Table(displayName=f"tbl_dict_{category_id}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(table)


def checks_sheet(wb: Workbook, category_id: str) -> None:
    ws = wb.create_sheet("Kiem tra")
    style_title(ws, "Kiem tra chat luong", "Checklist tu dong de phat hien dong thieu du lieu, canh bao va bat thuong.", 6)
    columns = SHEET_COLUMNS[category_id]
    status_col = next((i + 1 for i, name in enumerate(columns) if name == "Trang thai"), 8)
    owner_col = next((i + 1 for i, name in enumerate(columns) if "Phu trach" in name or "Nguoi phu trach" in name), 9)
    status_letter = get_column_letter(status_col)
    owner_letter = get_column_letter(owner_col)
    checks = [
        ("So dong co du lieu", f"=COUNTA('Nhap lieu'!A{DATA_START_ROW}:A{DATA_END_ROW})", "Theo doi quy mo du lieu"),
        ("Dong thieu trang thai", f"=COUNTBLANK('Nhap lieu'!{status_letter}{DATA_START_ROW}:{status_letter}{DATA_END_ROW})", "Can bo sung de dashboard dung"),
        ("Dong thieu nguoi phu trach", f"=COUNTBLANK('Nhap lieu'!{owner_letter}{DATA_START_ROW}:{owner_letter}{DATA_END_ROW})", "Can bo sung owner"),
        ("Dong can chu y", f'=COUNTIF(\'Nhap lieu\'!{status_letter}{DATA_START_ROW}:{status_letter}{DATA_END_ROW},"Can chu y")', "Can xu ly uu tien"),
        ("Gia tri am cot ke hoach", f"=COUNTIF('Nhap lieu'!D{DATA_START_ROW}:D{DATA_END_ROW},\"<0\")", "Khong nen co so am"),
        ("Gia tri am cot thuc hien", f"=COUNTIF('Nhap lieu'!E{DATA_START_ROW}:E{DATA_END_ROW},\"<0\")", "Khong nen co so am"),
    ]
    ws.append([])
    ws.append(["Hang muc kiem tra", "Ket qua", "Ghi chu"])
    for row in checks:
        ws.append(list(row))
    for cell in ws[4][0:3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.conditional_formatting.add("B6:B10", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FEE2E2")))
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 54


def executive_report_sheet(wb: Workbook, template: dict) -> None:
    ws = wb.create_sheet("Bao cao dieu hanh")
    ws.sheet_view.showGridLines = False
    style_title(ws, "Bao cao dieu hanh", "Trang tom tat co the copy sang email, Word hoac PowerPoint.", 6)
    sections = [
        ("Tom tat", f"Template {template['name']} da san sang cho nhap lieu, theo doi KPI va kiem tra chat luong."),
        ("Diem noi bat", "Co dashboard, bang du lieu co validation, chart, checklist loi va tu dien du lieu."),
        ("Hanh dong tiep theo", "Cap nhat sheet Nhap lieu, kiem tra tab Kiem tra, sau do dung Dashboard/Bao cao dieu hanh."),
        ("Rui ro can theo doi", "Dong co trang thai Can chu y, thieu owner, thieu trang thai hoac gia tri am."),
    ]
    row = 4
    for title, body in sections:
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="107C41")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        ws.cell(row=row, column=1, value=body)
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=6)
        row += 3
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18


def workbook_bytes(template: dict, index: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Nhap lieu"
    category_id = template["category_id"]
    style_data_sheet(ws, template["name"], category_id)
    add_formula_columns(ws, category_id)

    dashboard_sheet(wb, template, category_id)
    control_panel_sheet(wb, template)
    data_dictionary_sheet(wb, category_id)
    checks_sheet(wb, category_id)
    executive_report_sheet(wb, template)

    lists = wb.create_sheet("Danh muc")
    style_title(lists, "Danh muc dung chung", "Dung cho dropdown, chuan hoa trang thai va nguoi phu trach.", 5)
    lists.append([])
    lists.append(["Trang thai", "Nguoi phu trach", "Bo phan", "Muc do", "Ghi chu"])
    for idx in range(12):
        lists.append([
            STATUS_VALUES[idx % len(STATUS_VALUES)],
            OWNERS[idx % len(OWNERS)],
            DEPARTMENTS[idx % len(DEPARTMENTS)],
            ["Thap", "Trung binh", "Cao"][idx % 3],
            "Co the thay doi theo doanh nghiep",
        ])
    for cell in lists[4][0:5]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for col in range(1, 6):
        lists.column_dimensions[get_column_letter(col)].width = 22

    guide = wb.create_sheet("Huong dan")
    style_title(guide, "Huong dan su dung template", "Quy trinh de ap dung nhanh trong ExcelAI va Excel.", 5)
    guide_rows = [
        ("0", "Control Panel", "Dien tham so quan tri, nguong canh bao va nguoi so huu."),
        ("1", "Nhap lieu", "Cap nhat cac dong trong sheet Nhap lieu, khong xoa dong header."),
        ("2", "Kiem tra", "Dung filter va conditional formatting de tim dong can chu y."),
        ("3", "Dashboard", "Dashboard tu cap nhat bang cong thuc va bieu do."),
        ("4", "Tu dien du lieu", "Doc mo ta cot de map du lieu doanh nghiep vao template."),
        ("5", "Bao cao dieu hanh", "Copy noi dung tom tat sang email, Word hoac PowerPoint."),
        ("6", "Xuat bao cao", "Tai len ExcelAI de chat, lam sach, doi soat hoac tao bao cao."),
    ]
    for col, value in enumerate(["Buoc", "Khu vuc", "Mo ta"], start=1):
        guide.cell(row=4, column=col, value=value)
    for idx, row in enumerate(guide_rows, start=5):
        guide.cell(idx, 1, row[0])
        guide.cell(idx, 2, row[1])
        guide.cell(idx, 3, row[2])
    for cell in guide[4][0:3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    guide.column_dimensions["A"].width = 10
    guide.column_dimensions["B"].width = 20
    guide.column_dimensions["C"].width = 82

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.freeze_panes = sheet.freeze_panes or "A4"

    wb.active = 0
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def write_templates_to_filesystem(root: Path) -> int:
    templates = build_catalog()
    for index, template in enumerate(templates, start=1):
        folder = root / template["id"]
        folder.mkdir(parents=True, exist_ok=True)
        (folder / template["file"]).write_bytes(workbook_bytes(template, index))
        (folder / "preview.svg").write_bytes(preview_svg_bytes(template, index))
    return len(templates)


def seed_with_db() -> None:
    db = get_db()
    storage = StorageService(db)
    metadata = get_metadata(db)
    metadata.setdefault("templates", {})

    templates = build_catalog()
    for index, template in enumerate(templates, start=1):
        storage_path = f"templates/catalog/{template['id']}/{template['file']}"
        preview_path = f"templates/catalog/{template['id']}/preview.svg"
        storage.upload_bytes(
            storage_path,
            workbook_bytes(template, index),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        storage.upload_bytes(preview_path, preview_svg_bytes(template, index), "image/svg+xml")
        row = {
            "id": template["id"],
            "name": template["name"],
            "category": template["category"],
            "description": template["description"],
            "file": template["file"],
            "icon": template["icon"],
            "color": template["color"],
            "created_at": template["created_at"].isoformat(),
        }
        db.table("templates").upsert(row).execute()
        metadata["templates"][template["id"]] = {
            "storagePath": storage_path,
            "previewPath": preview_path,
            "previewContentType": "image/svg+xml",
            "fileName": template["file"],
            "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "seededBy": Path(__file__).name,
        }

    save_metadata(db, metadata)
    print(f"Seeded {len(templates)} detailed Excel templates into {settings.storage_bucket}.")


def main() -> None:
    if "--filesystem-only" in sys.argv:
        roots = [
            BACKEND_DIR / "storage" / settings.storage_bucket / "templates" / "catalog",
            BACKEND_DIR.parent / "storage" / settings.storage_bucket / "templates" / "catalog",
        ]
        total = 0
        for root in roots:
            total = write_templates_to_filesystem(root)
        print(f"Wrote {total} detailed Excel templates to {len(roots)} catalog folders.")
        return

    seed_with_db()


def preview_svg_bytes(template: dict, index: int) -> bytes:
    colors = {
        "green": ("#107C41", "#E7F6EE"),
        "blue": ("#2563EB", "#EAF1FF"),
        "purple": ("#7C3AED", "#F1EAFF"),
        "orange": ("#EA580C", "#FFF1E7"),
        "teal": ("#0F766E", "#E6F7F5"),
        "pink": ("#DB2777", "#FCE7F3"),
        "slate": ("#475569", "#EEF2F7"),
        "indigo": ("#4F46E5", "#EEF2FF"),
        "cyan": ("#0891B2", "#E6F9FE"),
        "amber": ("#D97706", "#FFF7E6"),
    }
    primary, soft = colors.get(template["color"], ("#107C41", "#E7F6EE"))
    columns = SHEET_COLUMNS[template["category_id"]]
    title = escape_xml(template["name"])
    category = escape_xml(template["category"])
    rows = []
    for row_index in range(5):
        y = 178 + row_index * 32
        opacity = 0.95 - row_index * 0.08
        rows.append(
            f'<rect x="58" y="{y}" width="524" height="24" rx="4" fill="#FFFFFF" opacity="{opacity:.2f}"/>'
            f'<rect x="72" y="{y + 7}" width="{120 + row_index * 8}" height="6" rx="3" fill="{primary}" opacity="0.45"/>'
            f'<rect x="246" y="{y + 7}" width="{96 + row_index * 5}" height="6" rx="3" fill="#94A3B8" opacity="0.65"/>'
            f'<rect x="402" y="{y + 7}" width="{72 + row_index * 4}" height="6" rx="3" fill="#CBD5E1" opacity="0.85"/>'
        )
    headers = "".join(
        f'<text x="{72 + i * 86}" y="154" font-family="Arial" font-size="11" font-weight="700" fill="#0F172A">{escape_xml(col[:10])}</text>'
        for i, col in enumerate(columns[:6])
    )
    svg = f'''<svg xmlns="http://www.w3.org/2000/svg" width="640" height="360" viewBox="0 0 640 360">
<rect width="640" height="360" rx="22" fill="#0B1220"/>
<rect x="28" y="28" width="584" height="304" rx="18" fill="{soft}"/>
<rect x="28" y="28" width="584" height="84" rx="18" fill="{primary}"/>
<text x="58" y="72" font-family="Arial" font-size="24" font-weight="700" fill="#FFFFFF">{title}</text>
<text x="58" y="96" font-family="Arial" font-size="13" fill="#FFFFFF" opacity="0.86">{category} · ExcelAI template #{index:03d}</text>
<rect x="58" y="128" width="524" height="190" rx="10" fill="#F8FAFC" stroke="#CBD5E1"/>
<rect x="58" y="128" width="524" height="36" rx="10" fill="#E2E8F0"/>
{headers}
{''.join(rows)}
<rect x="434" y="46" width="148" height="38" rx="10" fill="#FFFFFF" opacity="0.18"/>
<text x="458" y="71" font-family="Arial" font-size="13" font-weight="700" fill="#FFFFFF">XLSX READY</text>
</svg>'''
    return svg.encode("utf-8")


def escape_xml(value: str) -> str:
    return (
        str(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def get_metadata(db) -> dict:
    rows = db.table("settings").select("*").eq("key", "template_admin_metadata").limit(1).execute().data or []
    if not rows:
        return {"templates": {}}
    try:
        payload = json.loads(rows[0].get("value") or "{}")
        return payload if isinstance(payload, dict) else {"templates": {}}
    except json.JSONDecodeError:
        return {"templates": {}}


def save_metadata(db, metadata: dict) -> None:
    payload = {"key": "template_admin_metadata", "value": json.dumps(metadata, ensure_ascii=False)}
    db.table("settings").upsert(payload).execute()


if __name__ == "__main__":
    main()
