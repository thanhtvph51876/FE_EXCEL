from io import BytesIO
import json
import re

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

from dependencies import get_db
from services.storage_service import StorageService


router = APIRouter(prefix="/api/templates", tags=["templates"])


def _template_payload(row: dict) -> dict:
    return {
        "id": row.get("id"),
        "name": row.get("name"),
        "category": row.get("category") or "",
        "description": row.get("description") or "",
        "file": row.get("file") or "",
        "icon": row.get("icon") or "",
        "color": row.get("color") or "accent",
    }


def _safe_filename(value: str) -> str:
    base = re.sub(r"[^A-Za-z0-9_.-]+", "-", value or "excelai-template").strip(".-")
    if not base:
        base = "excelai-template"
    if not base.lower().endswith((".xlsx", ".xls", ".xlsm", ".csv")):
        base += ".xlsx"
    return base[:120]


def _build_template_workbook(template: dict) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Template"
    name = template.get("name") or "ExcelAI Template"
    category = template.get("category") or "General"
    description = template.get("description") or ""

    sheet["A1"] = name
    sheet["A2"] = f"Category: {category}"
    sheet["A3"] = description
    sheet["A5"] = "STT"
    sheet["B5"] = "Tên hạng mục"
    sheet["C5"] = "Giá trị"
    sheet["D5"] = "Ghi chú"
    for index in range(1, 6):
        row = index + 5
        sheet.cell(row=row, column=1, value=index)
        sheet.cell(row=row, column=2, value=f"Dòng mẫu {index}")
        sheet.cell(row=row, column=3, value=0)
        sheet.cell(row=row, column=4, value="")
    sheet.freeze_panes = "A6"
    for column, width in {"A": 10, "B": 28, "C": 18, "D": 36}.items():
        sheet.column_dimensions[column].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _template_metadata(db) -> dict:
    response = db.table("settings").select("*").eq("key", "template_admin_metadata").limit(1).execute()
    raw = response.data[0].get("value") if response.data else ""
    try:
        parsed = json.loads(raw) if raw else {}
        return (parsed.get("templates") or {}) if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        return {}


@router.get("")
async def list_templates(db = Depends(get_db)):
    response = db.table("templates").select("*").order("created_at", desc=True).execute()
    return {"templates": [_template_payload(row) for row in (response.data or [])]}


@router.get("/{template_id}")
async def get_template(template_id: str, db = Depends(get_db)):
    response = db.table("templates").select("*").eq("id", template_id).limit(1).execute()
    if not response.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Biểu mẫu không tồn tại.")
    return {"template": _template_payload(response.data[0])}


@router.get("/{template_id}/download")
async def download_template(template_id: str, db = Depends(get_db)):
    response = db.table("templates").select("*").eq("id", template_id).limit(1).execute()
    if not response.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Biểu mẫu không tồn tại.")
    template = _template_payload(response.data[0])
    metadata = _template_metadata(db).get(template_id) or {}
    if metadata.get("storagePath"):
        content = StorageService(db).download_bytes(metadata["storagePath"])
        file_name = _safe_filename(metadata.get("fileName") or template.get("file") or template.get("name") or template_id)
        return StreamingResponse(
            BytesIO(content),
            media_type="application/octet-stream",
            headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
        )
    file_name = _safe_filename(template.get("file") or template.get("name") or template_id)
    return StreamingResponse(
        _build_template_workbook(template),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )
